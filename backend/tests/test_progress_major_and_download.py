from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from app.api.routes.datasets import _generated_dataset_workbook
from app.core.config import get_settings
from app.db import Base
from app.models import DatasetVersion, Major
from app.models.progress_models import AssignmentType, ProgressAssignment
from app.services.dataset_service import get_active_dataset
from app.services.progress_processing import process_progress_report, read_progress_report
from app.services.progress_service import generate_report, preview_progress_upload, upload_progress_report
from app.services.progress_service import push_progress_to_advising
from app.services.storage import StorageService


def _xlsx_bytes(df: pd.DataFrame, sheet_name: str = 'Progress Report') -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    return buf.getvalue()


def test_generated_dataset_workbook_uses_parsed_payload():
    version = DatasetVersion(
        dataset_type='courses',
        version_label='abc123',
        original_filename='missing.xlsx',
        storage_key='missing',
        parsed_payload={'records': [{'ID': '1', 'NAME': 'Alice'}]},
        metadata_json={},
    )

    workbook_bytes = _generated_dataset_workbook(version)
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes))
    ws = wb['courses']

    assert ws['A1'].value == 'ID'
    assert ws['B1'].value == 'NAME'
    assert ws['A2'].value == '1'
    assert ws['B2'].value == 'Alice'


def test_progress_report_long_format_preserves_major():
    content = _xlsx_bytes(pd.DataFrame([
        {'ID': '1', 'NAME': 'Alice', 'MAJOR': 'Engineering', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
    ]))

    df = read_progress_report(content, 'progress.xlsx')

    assert df.loc[0, 'MAJOR'] == 'Engineering'


def test_progress_report_wide_format_preserves_major():
    content = _xlsx_bytes(pd.DataFrame([
        {'ID': '1', 'NAME': 'Alice', 'MAJOR': 'Engineering', 'COURSE_1': 'PBHL201/Fall-2025/A'},
    ]))

    df = read_progress_report(content, 'progress.xlsx')

    assert df.loc[0, 'MAJOR'] == 'Engineering'


def test_process_progress_report_carries_major_to_pivots():
    df = pd.DataFrame([
        {'ID': '1', 'NAME': 'Alice', 'MAJOR': 'Engineering', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
    ])

    req_df, int_df, _extra_df, _extra_list = process_progress_report(
        df,
        target_courses={'PBHL201': 3},
        intensive_courses={'ARAB101': 3},
        target_rules={'PBHL201': [{'Credits': 3, 'PassingGrades': 'A,B,C,CR', 'FromOrd': -1e9, 'ToOrd': 1e9}]},
        intensive_rules={'ARAB101': [{'Credits': 3, 'PassingGrades': 'A,B,C,CR', 'FromOrd': -1e9, 'ToOrd': 1e9}]},
    )

    assert req_df.loc[0, 'MAJOR'] == 'Engineering'
    assert int_df.loc[0, 'MAJOR'] == 'Engineering'


def test_push_progress_to_advising_preserves_major(tmp_path: Path, monkeypatch):
    monkeypatch.setenv('LOCAL_STORAGE_PATH', str(tmp_path / 'storage'))
    get_settings.cache_clear()

    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    session = Session()
    try:
        major = Major(code='TEST', name='Test Major')
        session.add(major)
        session.flush()
        session.add(
            DatasetVersion(
                major_id=major.id,
                dataset_type='course_config',
                version_label='config',
                is_active=True,
                parsed_payload={'records': [{
                    'target_courses': {'PBHL201': 3},
                    'intensive_courses': {},
                    'target_rules': {'PBHL201': [{'Credits': 3, 'PassingGrades': 'A,B,C,CR', 'FromOrd': -1e9, 'ToOrd': 1e9}]},
                    'intensive_rules': {},
                }]},
                metadata_json={},
            )
        )
        session.add(
            DatasetVersion(
                major_id=major.id,
                dataset_type='progress_report',
                version_label='progress',
                is_active=True,
                parsed_payload={'records': [{
                    'ID': '1',
                    'NAME': 'Alice',
                    'MAJOR': 'Engineering',
                    'Course': 'PBHL201',
                    'Grade': 'A',
                    'Year': 2025,
                    'Semester': 'Fall',
                }]},
                metadata_json={},
            )
        )
        session.commit()

        push_progress_to_advising(session, 'TEST', user_id=None)
        pushed = get_active_dataset(session, 'TEST', 'progress')
        assert pushed is not None
        records = pushed.parsed_payload['records']
        assert records[0]['MAJOR'] == 'Engineering'
    finally:
        session.close()
        get_settings.cache_clear()


def test_generate_report_returns_per_student_extra_courses(tmp_path: Path):
    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    session = Session()
    try:
        major = Major(code='TEST', name='Test Major')
        session.add(major)
        session.flush()
        session.add(
            DatasetVersion(
                major_id=major.id,
                dataset_type='course_config',
                version_label='config',
                is_active=True,
                parsed_payload={'records': [{
                    'target_courses': {'PBHL201': 3},
                    'intensive_courses': {},
                    'target_rules': {'PBHL201': [{'Credits': 3, 'PassingGrades': 'A,B,C,CR', 'FromOrd': -1e9, 'ToOrd': 1e9}]},
                    'intensive_rules': {},
                }]},
                metadata_json={},
            )
        )
        session.add(
            DatasetVersion(
                major_id=major.id,
                dataset_type='progress_report',
                version_label='progress',
                is_active=True,
                parsed_payload={'records': [
                    {'ID': '1', 'NAME': 'Alice', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
                    {'ID': '1', 'NAME': 'Alice', 'Course': 'PBHL450', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
                    {'ID': '2', 'NAME': 'Bob', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
                ]},
                metadata_json={},
            )
        )
        session.commit()

        report = generate_report(session, 'TEST', page_size=50)
        by_student = {row.student_id: row for row in report.required}

        assert report.extra_courses == ['PBHL450']
        assert by_student['1'].extra_courses == ['PBHL450']
        assert by_student['2'].extra_courses == []
    finally:
        session.close()


def test_generate_report_removes_assigned_extra_courses(tmp_path: Path):
    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    session = Session()
    try:
        major = Major(code='TEST', name='Test Major')
        session.add(major)
        session.flush()
        session.add(
            DatasetVersion(
                major_id=major.id,
                dataset_type='course_config',
                version_label='config',
                is_active=True,
                parsed_payload={'records': [{
                    'target_courses': {'PBHL201': 3},
                    'intensive_courses': {},
                    'target_rules': {'PBHL201': [{'Credits': 3, 'PassingGrades': 'A,B,C,CR', 'FromOrd': -1e9, 'ToOrd': 1e9}]},
                    'intensive_rules': {},
                }]},
                metadata_json={},
            )
        )
        session.add(
            DatasetVersion(
                major_id=major.id,
                dataset_type='progress_report',
                version_label='progress',
                is_active=True,
                parsed_payload={'records': [
                    {'ID': '1', 'NAME': 'Alice', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
                    {'ID': '1', 'NAME': 'Alice', 'Course': 'PBHL450', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
                ]},
                metadata_json={},
            )
        )
        session.add(AssignmentType(major_id=major.id, label='SCE', sort_order=0))
        session.add(ProgressAssignment(major_id=major.id, student_id='1', assignment_type='SCE', course_code='PBHL450'))
        session.commit()

        report = generate_report(session, 'TEST', page_size=50)
        assert report.extra_courses == []
        assert report.required[0].extra_courses == []
    finally:
        session.close()


def test_generate_report_recovers_major_from_stored_file_when_payload_is_legacy(tmp_path: Path, monkeypatch):
    monkeypatch.setenv('LOCAL_STORAGE_PATH', str(tmp_path / 'storage'))
    get_settings.cache_clear()

    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    content = _xlsx_bytes(pd.DataFrame([
        {'ID': '1', 'NAME': 'Alice', 'MAJOR': 'Engineering', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
    ]))
    storage_key = 'datasets/TEST/progress_report/source.xlsx'
    StorageService().put_bytes(storage_key, content)

    session = Session()
    try:
        major = Major(code='TEST', name='Test Major')
        session.add(major)
        session.flush()
        session.add(
            DatasetVersion(
                major_id=major.id,
                dataset_type='course_config',
                version_label='config',
                is_active=True,
                parsed_payload={'records': [{
                    'target_courses': {'PBHL201': 3},
                    'intensive_courses': {},
                    'target_rules': {'PBHL201': [{'Credits': 3, 'PassingGrades': 'A,B,C,CR', 'FromOrd': -1e9, 'ToOrd': 1e9}]},
                    'intensive_rules': {},
                }]},
                metadata_json={},
            )
        )
        session.add(
            DatasetVersion(
                major_id=major.id,
                dataset_type='progress_report',
                version_label='legacy',
                original_filename='source.xlsx',
                storage_key=storage_key,
                is_active=True,
                parsed_payload={'records': [
                    {'ID': '1', 'NAME': 'Alice', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
                ]},
                metadata_json={},
            )
        )
        session.commit()

        report = generate_report(session, 'TEST', page_size=50)
        assert report.required[0].major == 'Engineering'
    finally:
        session.close()
        get_settings.cache_clear()


def test_preview_progress_upload_reports_source_major_options(tmp_path: Path):
    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    content = _xlsx_bytes(pd.DataFrame([
        {'ID': '1', 'NAME': 'Alice', 'MAJOR': 'PBHL', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
        {'ID': '2', 'NAME': 'Bob', 'MAJOR': 'SPETHE', 'Course': 'SPTH201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
        {'ID': '2', 'NAME': 'Bob', 'MAJOR': 'SPETHE', 'Course': 'SPTH202', 'Grade': 'B', 'Year': 2025, 'Semester': 'Fall'},
    ]))

    session = Session()
    try:
        result = preview_progress_upload(session, 'TEST', content)
        assert result['requires_major_selection'] is True
        assert result['major_options'] == [
            {'major': 'PBHL', 'student_count': 1, 'row_count': 1},
            {'major': 'SPETHE', 'student_count': 1, 'row_count': 2},
        ]
    finally:
        session.close()


def test_upload_progress_report_filters_by_selected_source_major(tmp_path: Path, monkeypatch):
    monkeypatch.setenv('LOCAL_STORAGE_PATH', str(tmp_path / 'storage'))
    get_settings.cache_clear()

    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    content = _xlsx_bytes(pd.DataFrame([
        {'ID': '1', 'NAME': 'Alice', 'MAJOR': 'PBHL', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
        {'ID': '2', 'NAME': 'Bob', 'MAJOR': 'SPETHE', 'Course': 'SPTH201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
        {'ID': '2', 'NAME': 'Bob', 'MAJOR': 'SPETHE', 'Course': 'SPTH202', 'Grade': 'B', 'Year': 2025, 'Semester': 'Fall'},
    ]))

    session = Session()
    try:
        major = Major(code='TEST', name='Test Major')
        session.add(major)
        session.commit()

        result = upload_progress_report(session, 'TEST', 'source.xlsx', content, user_id=1, source_majors=['spethe'])
        version = get_active_dataset(session, 'TEST', 'progress_report')
        records = version.parsed_payload['records']

        assert result == {'student_count': 1, 'row_count': 2}
        assert {record['MAJOR'] for record in records} == {'SPETHE'}
        assert version.metadata_json['selected_source_majors'] == ['SPETHE']
        assert version.metadata_json['pre_filter_student_count'] == 2
        assert version.metadata_json['post_filter_student_count'] == 1
        assert version.metadata_json['post_filter_row_count'] == 2
    finally:
        session.close()
        get_settings.cache_clear()


def test_progress_upload_filters_by_source_major_and_cohort_year(tmp_path: Path, monkeypatch):
    monkeypatch.setenv('LOCAL_STORAGE_PATH', str(tmp_path / 'storage'))
    get_settings.cache_clear()

    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    content = _xlsx_bytes(pd.DataFrame([
        {'ID': '20250001', 'NAME': 'Alice', 'MAJOR': 'SPTH', 'Course': 'SPTH201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
        {'ID': '20260002', 'NAME': 'Bob', 'MAJOR': 'SPTH', 'Course': 'SPTH201', 'Grade': 'B', 'Year': 2026, 'Semester': 'Fall'},
        {'ID': '20250003', 'NAME': 'Cara', 'MAJOR': 'PBHL', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
    ]))

    session = Session()
    try:
        session.add(Major(code='TEST', name='Test Major'))
        session.commit()

        result = upload_progress_report(
            session,
            'TEST',
            'source.xlsx',
            content,
            user_id=1,
            source_majors=['SPTH'],
            cohort_years=['2025'],
        )
        version = get_active_dataset(session, 'TEST', 'progress_report')
        records = version.parsed_payload['records']

        assert result == {'student_count': 1, 'row_count': 1}
        assert [str(record['ID']) for record in records] == ['20250001']
        assert version.metadata_json['selected_source_majors'] == ['SPTH']
        assert version.metadata_json['selected_cohort_years'] == ['2025']
        assert version.metadata_json['cohort_options'] == [
            {'year': '2025', 'student_count': 2, 'row_count': 2},
            {'year': '2026', 'student_count': 1, 'row_count': 1},
        ]
        assert version.metadata_json['post_filter_student_count'] == 1
        assert version.metadata_json['post_filter_row_count'] == 1
    finally:
        session.close()
        get_settings.cache_clear()


def test_progress_preview_returns_sticky_filter_defaults_for_same_major(tmp_path: Path, monkeypatch):
    monkeypatch.setenv('LOCAL_STORAGE_PATH', str(tmp_path / 'storage'))
    get_settings.cache_clear()

    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    first_content = _xlsx_bytes(pd.DataFrame([
        {'ID': '20250001', 'NAME': 'Alice', 'MAJOR': 'SPTH', 'Course': 'SPTH201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
        {'ID': '20260002', 'NAME': 'Bob', 'MAJOR': 'SPTH', 'Course': 'SPTH201', 'Grade': 'B', 'Year': 2026, 'Semester': 'Fall'},
        {'ID': '20250003', 'NAME': 'Cara', 'MAJOR': 'PBHL', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
    ]))
    next_content = _xlsx_bytes(pd.DataFrame([
        {'ID': '20250004', 'NAME': 'Dana', 'MAJOR': 'SPTH', 'Course': 'SPTH202', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
        {'ID': '20260005', 'NAME': 'Eli', 'MAJOR': 'SPTH', 'Course': 'SPTH202', 'Grade': 'A', 'Year': 2026, 'Semester': 'Fall'},
        {'ID': '20250006', 'NAME': 'Fay', 'MAJOR': 'PBHL', 'Course': 'PBHL202', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
    ]))

    session = Session()
    try:
        session.add(Major(code='TEST', name='Test Major'))
        session.add(Major(code='OTHER', name='Other Major'))
        session.commit()

        upload_progress_report(
            session,
            'TEST',
            'source.xlsx',
            first_content,
            user_id=1,
            source_majors=['SPTH'],
            cohort_years=['2025'],
        )

        result = preview_progress_upload(session, 'TEST', next_content)
        other_result = preview_progress_upload(session, 'OTHER', next_content)

        assert result['default_source_majors'] == ['SPTH']
        assert result['default_cohort_years'] == ['2025']
        assert result['total_students'] == 1
        assert result['total_rows'] == 1
        assert other_result['default_source_majors'] == []
        assert other_result['default_cohort_years'] == []
        assert other_result['total_students'] == 3
    finally:
        session.close()
        get_settings.cache_clear()


def test_progress_preview_omits_unavailable_sticky_defaults(tmp_path: Path, monkeypatch):
    monkeypatch.setenv('LOCAL_STORAGE_PATH', str(tmp_path / 'storage'))
    get_settings.cache_clear()

    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    first_content = _xlsx_bytes(pd.DataFrame([
        {'ID': '20250001', 'NAME': 'Alice', 'MAJOR': 'SPTH', 'Course': 'SPTH201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
    ]))
    next_content = _xlsx_bytes(pd.DataFrame([
        {'ID': '20260002', 'NAME': 'Bob', 'MAJOR': 'PBHL', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2026, 'Semester': 'Fall'},
    ]))

    session = Session()
    try:
        session.add(Major(code='TEST', name='Test Major'))
        session.commit()

        upload_progress_report(
            session,
            'TEST',
            'source.xlsx',
            first_content,
            user_id=1,
            source_majors=['SPTH'],
            cohort_years=['2025'],
        )

        result = preview_progress_upload(session, 'TEST', next_content)

        assert result['default_source_majors'] == []
        assert result['default_cohort_years'] == []
        assert result['total_students'] == 1
        assert result['total_rows'] == 1
    finally:
        session.close()
        get_settings.cache_clear()


def test_changed_progress_filter_selection_becomes_next_default(tmp_path: Path, monkeypatch):
    monkeypatch.setenv('LOCAL_STORAGE_PATH', str(tmp_path / 'storage'))
    get_settings.cache_clear()

    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    content = _xlsx_bytes(pd.DataFrame([
        {'ID': '20250001', 'NAME': 'Alice', 'MAJOR': 'SPTH', 'Course': 'SPTH201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
        {'ID': '20260002', 'NAME': 'Bob', 'MAJOR': 'SPTH', 'Course': 'SPTH201', 'Grade': 'B', 'Year': 2026, 'Semester': 'Fall'},
        {'ID': '20260003', 'NAME': 'Cara', 'MAJOR': 'PBHL', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2026, 'Semester': 'Fall'},
    ]))

    session = Session()
    try:
        session.add(Major(code='TEST', name='Test Major'))
        session.commit()

        upload_progress_report(
            session,
            'TEST',
            'source.xlsx',
            content,
            user_id=1,
            source_majors=['SPTH'],
            cohort_years=['2025'],
        )
        upload_progress_report(
            session,
            'TEST',
            'source.xlsx',
            content,
            user_id=1,
            source_majors=['PBHL'],
            cohort_years=['2026'],
        )
        result = preview_progress_upload(session, 'TEST', content)
        assert result['default_source_majors'] == ['PBHL']
        assert result['default_cohort_years'] == ['2026']
    finally:
        session.close()
        get_settings.cache_clear()


def test_upload_progress_report_requires_source_major_when_major_column_exists(tmp_path: Path):
    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    content = _xlsx_bytes(pd.DataFrame([
        {'ID': '1', 'NAME': 'Alice', 'MAJOR': 'PBHL', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
    ]))

    session = Session()
    try:
        major = Major(code='TEST', name='Test Major')
        session.add(major)
        session.commit()

        try:
            upload_progress_report(session, 'TEST', 'source.xlsx', content, user_id=1)
        except ValueError as exc:
            assert 'Select at least one source major' in str(exc)
        else:
            raise AssertionError('Expected source major selection to be required.')
    finally:
        session.close()


def test_upload_progress_report_without_major_column_uploads_all_rows(tmp_path: Path, monkeypatch):
    monkeypatch.setenv('LOCAL_STORAGE_PATH', str(tmp_path / 'storage'))
    get_settings.cache_clear()

    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}", future=True)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    Base.metadata.create_all(bind=engine)

    content = _xlsx_bytes(pd.DataFrame([
        {'ID': '1', 'NAME': 'Alice', 'Course': 'PBHL201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
        {'ID': '2', 'NAME': 'Bob', 'Course': 'SPTH201', 'Grade': 'A', 'Year': 2025, 'Semester': 'Fall'},
    ]))

    session = Session()
    try:
        major = Major(code='TEST', name='Test Major')
        session.add(major)
        session.commit()

        result = upload_progress_report(session, 'TEST', 'legacy.xlsx', content, user_id=1)
        version = get_active_dataset(session, 'TEST', 'progress_report')

        assert result == {'student_count': 2, 'row_count': 2}
        assert len(version.parsed_payload['records']) == 2
        assert version.metadata_json['source_major_options'] == []
    finally:
        session.close()
        get_settings.cache_clear()
